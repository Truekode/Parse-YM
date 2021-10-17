from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.common.exceptions import NoSuchElementException 
from collections import OrderedDict
from datetime import datetime
from webdriver_manager.firefox import GeckoDriverManager

import openpyxl
import requests
import json
import time
import os

with open('brands.json') as f:
    links = json.load(f)

catalogs_links = []
tovars = []
tovars_full = []
options = Options()
options.headless = True
browser = webdriver.Firefox(options=options,executable_path=GeckoDriverManager().install())

def get_captcha():
    try:
        time.sleep(10)
        browser.find_element_by_css_selector("input.CheckboxCaptcha-Button").click()
        time.sleep(10)
        ready_page = "loading"
        while ready_page == 'loading':
            ready_page = browser.execute_script("return document.readyState")
    except NoSuchElementException:
        return False
    return True


def check_next():
    try:
        get_captcha()
        ready_page = "loading"
        while ready_page == 'loading':
            ready_page = browser.execute_script("return document.readyState")
        browser.find_element_by_css_selector("a._3OFYT")
    except NoSuchElementException:
        return False
    return True

def go_to_next():
    ready_page = "loading"
    while ready_page == 'loading':
        ready_page = browser.execute_script("return document.readyState")
    browser.find_element_by_css_selector("a._3OFYT").click()
    get_captcha()
    ready_page = "loading"
    while ready_page == 'loading':
        ready_page = browser.execute_script("return document.readyState")

def get_catalogs(brand_id):
    a_s = browser.find_elements_by_xpath("//a[@href]")
    for i in a_s:
        mystring = i.get_attribute("href")
        if brand_id in mystring:
            word = '/catalog--'
            if word in mystring: 
                catalogs_links.append(mystring)
    return catalogs_links

def delete_clon(tovarss):
    tavars_1 = []
    for link in tovarss:
        if 'reviews' not in link:
            tavars_1.append(link)
    y = list(OrderedDict.fromkeys(tavars_1))
    return y

def get_tovars(link_catalog, brand_id):
    browser.get(link_catalog)
    ready_page = "loading"
    while ready_page == 'loading':
        ready_page = browser.execute_script("return document.readyState")
    if (get_captcha()):
        if (check_next()):
            while check_next():
                browser.execute_script("window.scrollTo({top: (document.body.scrollHeight * 0.8), behavior: 'smooth'});")
                time.sleep(10)
                t_s = browser.find_elements_by_xpath("//a[@href]")
                for i in t_s:
                    mystring = i.get_attribute("href")
                    if brand_id in mystring:
                        word = '/product--'
                        if word in mystring: 
                            tovars.append(mystring)
                if (check_next()):
                    go_to_next()
                else:
                    break
        else:
            browser.execute_script("window.scrollTo({top: (document.body.scrollHeight * 0.8), behavior: 'smooth'});")
            time.sleep(10)
            t_s = browser.find_elements_by_xpath("//a[@href]")
            for i in t_s:
                mystring = i.get_attribute("href")
                if brand_id in mystring:
                    word = '/product--'
                    if word in mystring: 
                        tovars.append(mystring)
    else:
        if (check_next()):
            while check_next():
                browser.execute_script("window.scrollTo({top: (document.body.scrollHeight * 0.8), behavior: 'smooth'});")
                time.sleep(10)
                t_s = browser.find_elements_by_xpath("//a[@href]")
                for i in t_s:
                    mystring = i.get_attribute("href")
                    if brand_id in mystring:
                        word = '/product--'
                        if word in mystring: 
                            tovars.append(mystring)
                if (check_next()):
                    go_to_next()
                else:
                    break
        else:
            browser.execute_script("window.scrollTo({top: (document.body.scrollHeight * 0.8), behavior: 'smooth'});")
            time.sleep(10)
            t_s = browser.find_elements_by_xpath("//a[@href]")
            for i in t_s:
                mystring = i.get_attribute("href")
                if brand_id in mystring:
                    word = '/product--'
                    if word in mystring: 
                        tovars.append(mystring)

def check_element(css_selector):
    try:
        browser.find_element_by_css_selector(css_selector)
    except NoSuchElementException:
        return False
    return True
def check_elements(css_selector):
    try:
        browser.find_elements_by_css_selector(css_selector)
    except NoSuchElementException:
        return False
    return True

def get_file(url):
    response = requests.get(url, stream=True)
    return response

def save_data(name, file_data, brand):
    current_date = datetime.now().date()
    current_date = current_date.strftime('%d.%m.%Y')
    try:
        os.mkdir('img_' + brand)
    except Exception:
        bezpoleznaya_per = 0
    file = open('img_'+ brand + '/' + name, 'bw') #Бинарный режим, изображение передається байтами
    for chunk in file_data.iter_content(4096): # Записываем в файл по блочно данные
        file.write(chunk)
def get_name(url):
    name = url.split('/')[5]
    return name

def get_tovars_full(url, brand):
    tovar = {}
    link = url.split('&')
    for j in link:
        if 'sku' in j:
            tovar['articul'] = j.split('=')[1]
    if 'sku' not in url:
        tovar['articul'] = url.split('/')[4].split('?')[0]
    tovar['brand'] = brand
    characteristic = {}
    img_name = []
    img_link = []
    if (check_element('h1._2OAAC')):
        tovar['name'] = browser.find_element_by_css_selector("h1._2OAAC").text
    else:
        tovar['name'] = ''

    if (check_element('div._1uLae')):
        tovar['HTML'] = browser.find_element_by_css_selector("div._1uLae").get_attribute('outerHTML')
        tovar['about'] = browser.find_element_by_css_selector("div._1uLae").get_attribute('innerHTML')
    else:
        tovar['HTML'] = ''
        tovar['about'] = ''
        
    if (check_element('div._34FT3')):
        tovar['last_price'] = browser.find_element_by_css_selector("div._34FT3").text
    else:
        tovar['last_price'] = ''

    if (check_element('div._3kWlK')):
        tovar['price'] = browser.find_element_by_css_selector("div._3kWlK").text
    else:
        tovar['price'] = ''

    if (check_element('span.Vu-M2')):
        tovar['postav'] = browser.find_element_by_css_selector("span.Vu-M2").text
    else:
        tovar['postav'] = ''

    if (check_element('img._2gUfn')):
        for name in browser.find_elements_by_css_selector("img._2gUfn"):
            link = name.get_attribute("src")
            new_link = link.split('/')
            new_link[6] = 'orig'
            link = '/'.join(new_link)
            img_name.append(get_name(link))
            img_link.append(link)
            save_data(get_name(link),get_file(link),brand)
        tovar['images'] = img_name
        tovar['images_link'] = img_link
    else:
        if (check_element('img._3Wp6V')):
            link = browser.find_element_by_css_selector("img._3Wp6V").get_attribute("src")
            img_name.append(get_name(link))
            img_link.append(link)
            save_data(get_name(link),get_file(link),brand)
            tovar['images'] = img_name
            tovar['images_link'] = img_link
        else:
            tovar['images'] = []
            tovar['images_link'] = []

    tovar['link'] = url
    browser.execute_script("window.scrollTo({top: (document.body.scrollHeight * 0.5), behavior: 'smooth'});")
    time.sleep(10)
    ready_page = "loading"
    while ready_page == 'loading':
        ready_page = browser.execute_script("return document.readyState")
    if (check_element('a._1VmAF')):
        browser.find_element_by_css_selector("a._1VmAF").click()
        if(get_captcha()):
            ready_page = "loading"
            while ready_page == 'loading':
                ready_page = browser.execute_script("return document.readyState")
            if (check_element('._3PnEm')):
                for count, i in enumerate(browser.find_elements_by_css_selector("._3PnEm")):
                    characteristic[browser.find_elements_by_css_selector("div._2TxqA")[count].text] = i.text
                tovar['characts'] = characteristic
            else:
                tovar['characts'] = {}
        else:
            ready_page = "loading"
            while ready_page == 'loading':
                ready_page = browser.execute_script("return document.readyState")
            if (check_element('._3PnEm')):
                for count, i in enumerate(browser.find_elements_by_css_selector("._3PnEm")):
                    characteristic[browser.find_elements_by_css_selector("div._2TxqA")[count].text] = i.text
                tovar['characts'] = characteristic
            else:
                tovar['characts'] = {}
        browser.get(url)
        time.sleep(10)
        get_captcha()
    else:
        if (check_element('._3PnEm')):
            for count, i in enumerate(browser.find_elements_by_css_selector("._3PnEm")):
                characteristic[browser.find_elements_by_css_selector("div._2TxqA")[count].text] = i.text
            tovar['characts'] = characteristic
        else:
            tovar['characts'] = {}

    tovars_full.append(tovar)
    with open('test.json', 'w') as f:
        json.dump(tovars_full,f, ensure_ascii=False, indent=4)

def get_characters(url,brand):
    if(check_elements('._3I5WG')):
        print(len(browser.find_elements_by_css_selector('._3I5WG')[0].find_elements_by_css_selector('._27xuj')))
        for n in range(len(browser.find_elements_by_css_selector('._3I5WG')[0].find_elements_by_css_selector('._27xuj'))):
            browser.find_elements_by_css_selector('._3I5WG')[0].find_elements_by_css_selector('._27xuj')[n].click()
            get_captcha()
            ready_page = "loading"
            while ready_page == 'loading':
                ready_page = browser.execute_script("return document.readyState")
            time.sleep(10)
            try:
                get_tovars_full(url,brand)
            except Exception as e:
                print(e)
            get_captcha()
    else:
        try:
            get_tovars_full(url,brand)
        except Exception as e:
            print(e)

def to_excel(tovars):
    book = openpyxl.Workbook()

    sheet = book.active

    def check_index(arr, key):
        return arr.index(key)

    keys = []
    keys.append('Артикул')
    keys.append('Бренд')
    keys.append('Поставщик')
    keys.append('Имя товара')
    keys.append('Описание с HTML')
    keys.append('Описание')
    keys.append('Цена')
    keys.append('Старая цена')
    keys.append('Имена картинок')
    keys.append('Ссылки на картинки')
    keys.append('Ссылка на товар')
    for i in tovars:
        for key in i['characts'].keys():
            keys.append(key)
    y = list(OrderedDict.fromkeys(keys))

    colomns = 1

    for count,item in enumerate(y):
        sheet.cell(row = 1, column=colomns).value = item
        colomns = colomns +1

    rows = 2

    for count, item in enumerate(tovars):
        sheet.cell(row = rows, column=1).value = item['articul']
        sheet.cell(row = rows, column=2).value = item['brand']
        sheet.cell(row = rows, column=3).value = item['postav']
        sheet.cell(row = rows, column=4).value = item['name']
        sheet.cell(row = rows, column=5).value = item['HTML']
        sheet.cell(row = rows, column=6).value = item['about']
        sheet.cell(row = rows, column=7).value = item['price']
        sheet.cell(row = rows, column=8).value = item['last_price']
        sheet.cell(row = rows, column=9).value = ';'.join(item['images'])
        sheet.cell(row = rows, column=10).value = ';'.join(item['images_link'])
        sheet.cell(row = rows, column=11).value = item['link']
        for key,val in item['characts'].items():
            sheet.cell(row = rows, column=check_index(y, key)+1).value = val
        rows = rows + 1

    current_date = datetime.now().date().strftime('%d.%m.%Y')
    book.save("Excel/" + tovars[0]['brand'] + '_' + current_date +".xlsx")
    book.close()


try:
    for brand_item in links:
        print('Cбор бренда ' + brand_item['name'])
        catalogs_links = []
        tovars = []
        tovars_full = []
        browser.get(brand_item['Url'])
        if (get_captcha()):
            time.sleep(10)
            catalog = get_catalogs(brand_item['id'])
            for count, i in enumerate(catalog):
                tmp = count + 1
                print('Сборка каталогов бренда ' + str(tmp) + '/' + str(len(catalog)))
                try:
                    get_tovars(i, brand_item['id'])
                except Exception as e:
                    print(e)
            tovars1 = delete_clon(tovars)
            for count1, tov in enumerate(tovars1):
                browser.get(tov)
                if (get_captcha()):
                    time.sleep(10)
                    browser.execute_script("window.scrollTo({top: (document.body.scrollHeight * 0.8), behavior: 'smooth'});")
                    ready_page = "loading"
                    while ready_page == 'loading':
                        ready_page = browser.execute_script("return document.readyState")
                    try:
                        get_characters(tov,brand_item['name'])
                    except Exception as e:
                        print(e)         
                else:
                    try:
                        get_characters(tov,brand_item['name'])
                    except Exception as e:
                        print(e) 
                tmp = count1 + 1
                print('Сборка характеристик товаров бренда ' + str(tmp) + '/' + str(len(tovars1)))
        else:
            catalog = get_catalogs(brand_item['id'])
            for count, i in enumerate(catalog):
                tmp = count + 1
                print('Сборка каталогов бренда ' + str(tmp) + '/' + str(len(catalog)))
                try:
                    get_tovars(i, brand_item['id'])
                except Exception as e:
                    print(e)
            tovars1 = delete_clon(tovars)
            for count1, tov in enumerate(tovars1):
                browser.get(tov)
                if (get_captcha()):
                    time.sleep(10)
                    browser.execute_script("window.scrollTo({top: (document.body.scrollHeight * 0.8), behavior: 'smooth'});")
                    ready_page = "loading"
                    while ready_page == 'loading':
                        ready_page = browser.execute_script("return document.readyState")
                    try:
                        get_characters(tov,brand_item['name'])
                    except Exception as e:
                        print(e)
                else:
                    try:
                        get_characters(tov,brand_item['name'])
                    except Exception as e:
                        print(e)
                tmp = count1 + 1
                print('Сборка характеристик товаров бренда ' + str(tmp) + '/' + str(len(tovars1)))

        to_excel(tovars_full)
        print('Готово!')
        input('Нажмите любую клавишу, чтобы выйти ... ')
except Exception as e:
    print(e)

browser.quit()

